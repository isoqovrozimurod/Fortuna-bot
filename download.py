from __future__ import annotations

import asyncio
import base64
import csv
import io
import json
import logging
import os
from pathlib import Path

import gspread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram import Bot, F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials

load_dotenv()
logger = logging.getLogger(__name__)
router = Router()

ADMIN_ID = int(os.getenv("ADMIN_ID", "0"))
SPREADSHEET_ID = "1UU87w2q9zk8q5_3pQqfVhp0Zp2hnU70bWWgu1R9q3No"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

TEMP_DIR = Path(__file__).resolve().parent / "temp"

# ===================== YUKLAB OLINADIGAN ELEMENTLAR =====================
# Yangi element qo'shish uchun shu ro'yxatga yozing:
# {"id": "unique_id", "label": "Tugma matni", "type": "sheet"|"file"}

ITEMS = [
    {
        "id": "users_sheet",
        "label": "👥 Foydalanuvchilar",
        "type": "sheet",
        "sheet": "user",
        "filename": "foydalanuvchilar.csv",
    },
    {
        "id": "subadmin_sheet",
        "label": "🛡 Sub-adminlar",
        "type": "sheet",
        "sheet": "Sub-adminlar",
        "filename": "sub_adminlar.csv",
    },
    {
        "id": "branches_sheet",
        "label": "🏢 Filiallar",
        "type": "sheet",
        "sheet": "malumotlar",
        "filename": "filiallar.csv",
    },
    {
        "id": "promo_image",
        "label": "🖼 Promo rasm",
        "type": "file",
        "path": str(TEMP_DIR / "fortuna.jpg"),
        "filename": "fortuna.jpg",
    },
]


# ===================== SHEETS CLIENT =====================

_gc: gspread.Client | None = None


def _get_gc() -> gspread.Client:
    global _gc
    if _gc is None:
        b64 = os.getenv("GOOGLE_CREDENTIALS_B64")
        creds_dict = json.loads(base64.b64decode(b64).decode())
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        _gc = gspread.authorize(creds)
    return _gc


def _sheet_to_csv(sheet_name: str) -> bytes:
    gc = _get_gc()
    sh = gc.open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet(sheet_name)
    rows = ws.get_all_values()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # utf-8-sig — Excel uchun


# ===================== TOZALASH =====================

def _cleanup_sync(sheet_name: str) -> None:
    """
    Yuklanishdan avval sheet ni tozalaydi:
    - Telegram ID bo'sh qatorlarni o'chiradi
    - T/r ni 1 dan tartiblab qayta yozadi
    """
    gc = _get_gc()
    sh = gc.open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet(sheet_name)

    all_rows = ws.get_all_values()
    if len(all_rows) <= 1:
        return

    data_rows = all_rows[1:]

    # Telegram ID (2-ustun) bo'sh bo'lmagan qatorlarni olamiz
    valid_rows = [
        row for row in data_rows
        if len(row) > 1 and str(row[1]).strip()
    ]

    if not valid_rows:
        return

    # T/r ni 1 dan qayta tartiblаymiz
    for i, row in enumerate(valid_rows, start=1):
        while len(row) < 8:
            row.append("")
        row[0] = str(i)

    total_existing = len(all_rows)
    total_valid = len(valid_rows)

    # Mavjud ma'lumotlarni qayta yozamiz
    ws.update(
        f"A2:H{total_valid + 1}",
        valid_rows,
        value_input_option="RAW"
    )

    # Ortiqcha qatorlarni tozalaymiz
    if total_existing > total_valid + 1:
        empty_rows = [[""] * 8 for _ in range(total_existing - total_valid - 1)]
        ws.update(
            f"A{total_valid + 2}:H{total_existing}",
            empty_rows,
            value_input_option="RAW"
        )


# Tozalanadigan sheetlar ro'yxati (sheet nomi)
CLEANABLE_SHEETS = {"user", "Sub-adminlar"}


async def _cleanup_before_download(sheet_name: str) -> None:
    """Tozalanadigan sheet bo'lsa — avval tozalaydi"""
    if sheet_name not in CLEANABLE_SHEETS:
        return
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, _cleanup_sync, sheet_name)
    except Exception as e:
        logger.warning(f"Tozalashda xato ({sheet_name}): {e}")


# ===================== EXCEL EXPORT =====================

SHEET_ITEMS = [item for item in ITEMS if item["type"] == "sheet"]


def _all_to_excel() -> bytes:
    """Barcha Google Sheets varaqlarini bitta Excel faylga yozadi"""
    gc = _get_gc()
    sh = gc.open_by_key(SPREADSHEET_ID)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Bo'sh default sheetni o'chiramiz

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E7D32")
    center = Alignment(horizontal="center", vertical="center")

    for item in SHEET_ITEMS:
        ws_gsheet = sh.worksheet(item["sheet"])
        rows = ws_gsheet.get_all_values()
        if not rows:
            continue

        ws_excel = wb.create_sheet(title=item["label"].replace("📊", "").replace("👥", "").replace("🛡", "").replace("🏢", "").strip())

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws_excel.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center

        # Ustun kengligini avtomatik sozlaymiz
        for col in ws_excel.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws_excel.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===================== KLAVIATURA =====================

def download_menu_kb() -> InlineKeyboardMarkup:
    """Elementlarni 2 ta ustunda chiqaradi"""
    buttons = [
        InlineKeyboardButton(text=item["label"], callback_data=f"dl_{item['id']}")
        for item in ITEMS
    ]
    # 2 ta ustun
    rows = [buttons[i:i+2] for i in range(0, len(buttons), 2)]
    # Barchasini bitta fayl
    rows.append([InlineKeyboardButton(text="📦 Barchasini yuklab olish", callback_data="dl_all")])
    rows.append([InlineKeyboardButton(text="❌ Yopish", callback_data="dl_close")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ===================== HANDLERLAR =====================

@router.message(Command("download"))
async def cmd_download(message: Message):
    if message.from_user.id != ADMIN_ID:
        return
    await message.answer(
        "📥 <b>Yuklab olish</b>\n\nQaysi ma\'lumotni yuklab olmoqchisiz?",
        reply_markup=download_menu_kb(),
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("dl_"))
async def handle_download(call: CallbackQuery, bot: Bot):
    if call.from_user.id != ADMIN_ID:
        await call.answer("❌ Ruxsat yo'q!", show_alert=True)
        return

    item_id = call.data[3:]  # "dl_" ni olib tashlaymiz

    if item_id == "close":
        await call.message.delete()
        return

    if item_id == "all":
        await call.answer("⏳ Tayyorlanmoqda...")
        try:
            # Barcha tozalanadigan sheetlarni avval tozalaymiz
            for sheet_name in CLEANABLE_SHEETS:
                await _cleanup_before_download(sheet_name)
            loop = asyncio.get_event_loop()
            data = await loop.run_in_executor(None, _all_to_excel)
            file = BufferedInputFile(data, filename="fortuna_biznes_malumotlar.xlsx")
            await bot.send_document(
                call.from_user.id,
                file,
                caption="""📦 <b>Barcha ma\'lumotlar</b>
                         📁 fortuna_biznes_malumotlar.xlsx""",
                         parse_mode="HTML",
            )
        except Exception as e:
            logger.error(f"Excel export xato: {e}")
            await bot.send_message(
                call.from_user.id,
                f"❌ Xato: <code>{e}</code>",
                parse_mode="HTML",
            )
        return

    item = next((i for i in ITEMS if i["id"] == item_id), None)
    if not item:
        await call.answer("❌ Topilmadi!", show_alert=True)
        return

    await call.answer("⏳ Tayyorlanmoqda...")

    try:
        if item["type"] == "sheet":
            # Avval tozalaymiz (bo'sh qatorlar, T/r tartib)
            await _cleanup_before_download(item["sheet"])
            # Keyin yuklaymiz
            loop = asyncio.get_event_loop()
            data = await loop.run_in_executor(
                None, _sheet_to_csv, item["sheet"]
            )
            file = BufferedInputFile(data, filename=item["filename"])
            await bot.send_document(
                call.from_user.id,
                file,
                caption=f"📊 {item['label']}\n📁 {item['filename']}",
            )

        elif item["type"] == "file":
            path = Path(item["path"])
            if not path.exists():
                await call.answer("❌ Fayl topilmadi!", show_alert=True)
                return
            data = path.read_bytes()
            file = BufferedInputFile(data, filename=item["filename"])
            await bot.send_document(
                call.from_user.id,
                file,
                caption=f"📎 {item['label']}\n📁 {item['filename']}",
            )

    except Exception as e:
        logger.error(f"Yuklab olishda xato ({item_id}): {e}")
        await bot.send_message(
            call.from_user.id,
            f"❌ Xato yuz berdi:\n<code>{e}</code>",
            parse_mode="HTML",
        )
